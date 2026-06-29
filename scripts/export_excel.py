#!/usr/bin/env python3
"""导出 6月21-24日 邮箱收取记录和推送记录为两个 Excel 文件。"""
import sqlite3
import json
import sys
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DB_PATH = "/Users/admin/Desktop/邮箱抓取简历/data/processed.db"
OUTPUT_DIR = Path("/Users/admin/Desktop/邮箱抓取简历/exports")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

DATE_FROM = "2026-06-21"
DATE_TO = "2026-06-25"  # exclusive (covers 21, 22, 23, 24)

# ── Styles ──────────────────────────────────────────────
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_FONT = Font(name="微软雅黑", size=10)
CELL_ALIGN = Alignment(vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

TAG_FILLS = {
    "已匹配": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "未匹配": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "成功": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "失败": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "部分成功": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "推送中": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
}


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def style_cells(ws, nrows, ncols):
    for row in range(2, nrows + 2):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = CELL_FONT
            cell.alignment = CELL_ALIGN
            cell.border = THIN_BORDER


def auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def fmt_time(s):
    """ISO 时间字符串 → 可读格式"""
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return s


def export_email_records(conn):
    """导出邮箱收取记录"""
    sql = """
        SELECT account_name, subject, mail_date, processed_at, saved_files, matched
        FROM processed_emails
        WHERE processed_at >= ? AND processed_at < ?
        ORDER BY processed_at ASC
    """
    rows = conn.execute(sql, (DATE_FROM, DATE_TO)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "邮箱收取记录"

    headers = ["账号", "邮件主题", "邮件时间", "收取时间", "附件数", "状态"]
    ws.append(headers)

    for r in rows:
        saved_files = json.loads(r["saved_files"]) if r["saved_files"] else []
        matched_text = "已匹配" if r["matched"] else "未匹配"
        ws.append([
            r["account_name"],
            r["subject"] or "",
            fmt_time(r["mail_date"]),
            fmt_time(r["processed_at"]),
            len(saved_files),
            matched_text,
        ])

    style_header(ws, len(headers))
    style_cells(ws, len(rows), len(headers))
    auto_width(ws, [16, 50, 20, 20, 8, 10])

    # Color the status column
    for row in range(2, len(rows) + 2):
        val = ws.cell(row=row, column=6).value
        if val in TAG_FILLS:
            ws.cell(row=row, column=6).fill = TAG_FILLS[val]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows) + 1}"

    path = OUTPUT_DIR / "邮箱收取记录_20260621-20260624.xlsx"
    wb.save(str(path))
    print(f"✅ 邮箱收取记录: {len(rows)} 条 → {path}")
    return path


def parse_response_summary(response_body, error_message):
    """从 response_body JSON 提取推送结果摘要"""
    if not response_body:
        return {"created": None, "updated": None, "failed": None, "skipped": None, "summary": error_message or ""}

    try:
        data = json.loads(response_body)
    except (json.JSONDecodeError, TypeError):
        return {"created": None, "updated": None, "failed": None, "skipped": None, "summary": error_message or ""}

    # Success format: {"total":9,"created":7,"updated":2,"skipped":0,"failed":0,...}
    if "total" in data or "created" in data:
        created = data.get("created")
        updated = data.get("updated")
        failed = data.get("failed")
        skipped = data.get("skipped")
        parts = []
        if "total" in data:
            parts.append(f"合计{data['total']}")
        if created is not None:
            parts.append(f"新建{created}")
        if updated is not None:
            parts.append(f"更新{updated}")
        if failed:
            parts.append(f"失败{failed}")
        if skipped:
            parts.append(f"跳过{skipped}")
        failures = data.get("failures", [])
        if failures:
            parts.append(f"失败明细{len(failures)}条")
        summary = " · ".join(parts) if parts else ""
        return {"created": created, "updated": updated, "failed": failed, "skipped": skipped, "summary": summary}

    # Error format: {"type":"...","title":"...","status":422,"detail":"...","error_code":"..."}
    detail = data.get("detail") or data.get("title") or error_message or ""
    error_code = data.get("error_code", "")
    status_code = data.get("status", "")
    summary = f"{detail}"
    if error_code:
        summary += f"（{error_code}）"
    if status_code:
        summary = f"[{status_code}] {summary}"
    return {"created": None, "updated": None, "failed": None, "skipped": None, "summary": summary}


def export_push_records(conn):
    """导出推送记录"""
    sql = """
        SELECT id, excel_path, tenant_code, tenant_id, candidate_count,
               status, trigger_type, attempt_count, response_status,
               response_body, error_message, created_at, pushed_at
        FROM push_batches
        WHERE (pushed_at != '' AND pushed_at >= ? AND pushed_at < ?)
           OR (pushed_at = '' AND created_at >= ? AND created_at < ?)
        ORDER BY COALESCE(pushed_at, created_at) ASC
    """
    rows = conn.execute(sql, (DATE_FROM, DATE_TO, DATE_FROM, DATE_TO)).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "推送记录"

    headers = [
        "推送时间", "租户名", "推送人数", "新建", "更新",
        "失败", "跳过", "触发方式", "状态", "接口返回摘要"
    ]
    ws.append(headers)

    trigger_map = {"retry": "手动重推", "manual": "手动推送", "auto": "解析后自动"}
    status_map = {"success": "成功", "failed": "失败", "partial": "部分成功", "pushing": "推送中"}

    for r in rows:
        rs = parse_response_summary(r["response_body"], r["error_message"])
        pushed_time = r["pushed_at"] or r["created_at"]
        trigger = trigger_map.get(r["trigger_type"], r["trigger_type"])
        status_text = status_map.get(r["status"], r["status"])

        ws.append([
            fmt_time(pushed_time),
            r["tenant_code"] or "",
            r["candidate_count"] or 0,
            rs["created"] if rs["created"] is not None else "—",
            rs["updated"] if rs["updated"] is not None else "—",
            rs["failed"] if rs["failed"] is not None else "—",
            rs["skipped"] if rs["skipped"] is not None else "—",
            trigger,
            status_text,
            rs["summary"] or "—",
        ])

    style_header(ws, len(headers))
    style_cells(ws, len(rows), len(headers))
    auto_width(ws, [20, 12, 10, 8, 8, 8, 8, 12, 10, 50])

    # Color the status column (col 9)
    for row in range(2, len(rows) + 2):
        val = ws.cell(row=row, column=9).value
        if val in TAG_FILLS:
            ws.cell(row=row, column=9).fill = TAG_FILLS[val]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:J{len(rows) + 1}"

    path = OUTPUT_DIR / "推送记录_20260621-20260624.xlsx"
    wb.save(str(path))
    print(f"✅ 推送记录: {len(rows)} 条 → {path}")
    return path


def main():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    try:
        p1 = export_email_records(conn)
        p2 = export_push_records(conn)
        print(f"\n导出完成！文件位于: {OUTPUT_DIR}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
