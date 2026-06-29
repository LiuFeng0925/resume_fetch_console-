"""导出收取记录明细 + 解析记录明细（6月21日-24日）"""
from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl")
    raise

DB_PATH = Path(__file__).resolve().parent.parent / "data" / "processed.db"
EXPORT_DIR = Path(__file__).resolve().parent.parent / "exports"
EXPORT_DIR.mkdir(exist_ok=True)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TAG_OK = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
TAG_FAIL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for column_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length * 1.3, 10), 50)


def freeze(ws):
    ws.freeze_panes = "A2"


# ============================================================
# 1) 收取记录明细
# ============================================================
def export_fetch_records():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    rows = conn.execute(
        """
        SELECT account_name, subject, mail_date, processed_at,
               saved_files, matched
        FROM processed_emails
        WHERE processed_at >= '2026-06-21' AND processed_at < '2026-06-25'
        ORDER BY processed_at DESC
        """
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "收取记录"

    headers = ["邮箱", "邮件主题", "邮件时间", "收取时间", "附件数", "状态"]
    ws.append(headers)

    for r in rows:
        saved_files = r["saved_files"] or "[]"
        try:
            files = json.loads(saved_files) if isinstance(saved_files, str) else saved_files
            count = len(files) if isinstance(files, list) else 0
        except Exception:
            count = 0
        matched_str = "已匹配" if r["matched"] else "未匹配"
        ws.append([
            r["account_name"],
            r["subject"],
            r["mail_date"].replace("+08:00", "")[:19] if r["mail_date"] else "",
            r["processed_at"].replace("+08:00", "")[:19] if r["processed_at"] else "",
            count,
            matched_str,
        ])

    style_header(ws, len(headers))
    # 状态列着色
    status_col = 6
    for row in range(2, len(rows) + 2):
        cell = ws.cell(row=row, column=status_col)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if cell.value == "已匹配":
            cell.fill = TAG_OK
        elif cell.value == "未匹配":
            cell.fill = TAG_FAIL
    auto_width(ws)
    freeze(ws)

    path = EXPORT_DIR / "收取记录明细_20260621-20260624.xlsx"
    wb.save(path)
    print(f"✅ 收取记录明细: {path} ({len(rows)} 条)")
    return path


# ============================================================
# 2) 解析记录明细（从 parsed_json 提取候选人信息）
# ============================================================
def export_parse_records():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    rows = conn.execute(
        """
        SELECT source_file, file_path, account_name, job_id,
               tenant_id, tenant_code, status,
               parsed_json, used_ocr,
               excel_path, parsed_at,
               push_status, pushed_at, push_error, push_batch_id,
               candidate_name
        FROM parse_records
        WHERE parsed_at >= '2026-06-21' AND parsed_at < '2026-06-25'
        ORDER BY parsed_at DESC
        """
    ).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "解析记录"

    headers = [
        "文件名", "候选人姓名", "手机号", "邮箱编号",
        "租户ID", "租户编码",
        "解析状态", "是否OCR",
        "解析时间", "推送状态", "推送时间",
    ]
    ws.append(headers)

    for r in rows:
        # 从 parsed_json 中提取手机号
        phone = ""
        try:
            pj = json.loads(r["parsed_json"]) if isinstance(r["parsed_json"], str) else (r["parsed_json"] or {})
            if isinstance(pj, dict):
                phone = str(pj.get("phone") or "")
        except Exception:
            pass

        name = r["candidate_name"] or ""
        if not name and r["parsed_json"]:
            try:
                pj = json.loads(r["parsed_json"])
                name = pj.get("name") or ""
            except Exception:
                pass

        push_time = (r["pushed_at"] or "").replace("+08:00", "")[:19] if r["pushed_at"] else ""

        ws.append([
            r["source_file"],
            name,
            phone,
            r["job_id"],
            r["tenant_id"],
            r["tenant_code"],
            r["status"],
            "是" if r["used_ocr"] else "否",
            (r["parsed_at"] or "").replace("+08:00", "")[:19],
            r["push_status"] or "",
            push_time,
        ])

    style_header(ws, len(headers))

    # 解析状态着色
    status_col_map = {"成功": 7, "失败": 7}  # 解析状态在第7列
    for row in range(2, len(rows) + 2):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER
        sc = ws.cell(row=row, column=7)   # 解析状态
        pc = ws.cell(row=row, column=10)  # 推送状态
        if sc.value == "成功":
            sc.fill = TAG_OK
        elif sc.value in ("失败", "解析失败"):
            sc.fill = TAG_FAIL
        if pc.value == "success" or pc.value == "已推送":
            pc.fill = TAG_OK
        elif pc.value in ("failed", "skipped"):
            pc.fill = TAG_FAIL

    auto_width(ws)
    freeze(ws)

    path = EXPORT_DIR / "解析记录明细_20260621-20260624.xlsx"
    wb.save(path)
    print(f"✅ 解析记录明细: {path} ({len(rows)} 条)")
    return path


if __name__ == "__main__":
    p1 = export_fetch_records()
    p2 = export_parse_records()
    print(f"\n两个文件已保存到: {EXPORT_DIR}")
